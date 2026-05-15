"""Day 3: formula extraction (xlsx only).

Companion to extract.py. Opens every .xlsx release with openpyxl in NON-
read-only mode and data_only=False, captures formula strings, and writes
data/formulas/<release>.parquet.

xls files (2019 Q1-Q3, 2020 Q1) are SKIPPED — xlrd cannot expose Excel
formula strings (only opaque BIFF tokens). Coverage is therefore xlsx-only,
which means the audit cannot inspect formulas in the four .xls releases.

Schema:
    release_file, release_year, release_quarter
    sheet_name, sheet_idx
    row_idx, col_idx        (1-indexed, matches cells.parquet)
    formula                 the formula string with leading '='

Note: NON-read-only mode is slower and uses more memory than the value
extraction. We process one release at a time and discard immediately.

Cells with hardcoded values are NOT in this output — only cells whose
data_type indicates a formula. Joining to cells.parquet (LEFT JOIN on
release_file/sheet_name/row_idx/col_idx) gives the cached value alongside
the formula for any audit that needs both.
"""

from __future__ import annotations

import re
import sys
import time
from pathlib import Path

import openpyxl
import pandas as pd

SCRIPT_DIR = Path(__file__).parent
BRH_SOURCE = (SCRIPT_DIR / ".." / ".." / "FM Test" / "brh-dashboard"
              / "data" / "raw").resolve()
OUTPUT_DIR = SCRIPT_DIR / "data" / "formulas"

# Match brh_trim<N>_<YYYY>.xlsx. xls files explicitly excluded.
RELEASE_RE = re.compile(r"^brh_trim(\d)_(\d{4})\.xlsx$")
YEAR_MIN, YEAR_MAX = 2019, 2026


def discover_releases() -> list[tuple[Path, int, int]]:
    out = []
    for p in BRH_SOURCE.iterdir():
        m = RELEASE_RE.match(p.name)
        if not m:
            continue
        quarter, year = int(m.group(1)), int(m.group(2))
        if YEAR_MIN <= year <= YEAR_MAX:
            out.append((p, year, quarter))
    out.sort(key=lambda t: (t[1], t[2]))
    return out


def extract_release_formulas(path: Path, year: int, quarter: int) -> pd.DataFrame:
    """Open one xlsx release with formulas mode on, dump every formula cell."""
    rows: list[dict] = []
    # NON-read-only is required for formula access. data_only=False returns
    # formula STRINGS; data_only=True would return cached values.
    wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            try:
                ws = wb[sheet_name]
            except Exception as e:
                print(f"    ! sheet {sheet_idx:3d} {sheet_name!r}: "
                      f"failed to open ({type(e).__name__}: {e})", flush=True)
                continue
            # Iterate only used range; iter_rows is fastest.
            for row in ws.iter_rows():
                for cell in row:
                    # data_type 'f' indicates a formula in openpyxl.
                    # Some formulas may also be detected by string-prefix '='
                    # if data_type is unset.
                    if cell.data_type == 'f' or (
                        isinstance(cell.value, str) and cell.value.startswith('=')
                    ):
                        rows.append({
                            "release_file":    path.name,
                            "release_year":    year,
                            "release_quarter": quarter,
                            "sheet_name":      sheet_name,
                            "sheet_idx":       sheet_idx,
                            "row_idx":         cell.row,
                            "col_idx":         cell.column,
                            "formula":         str(cell.value),
                        })
    finally:
        wb.close()
    return pd.DataFrame(rows)


def write_parquet(df: pd.DataFrame, release_file: str) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if not df.empty:
        df = df.astype({
            "release_year":    "int16",
            "release_quarter": "int8",
            "sheet_idx":       "int16",
            "row_idx":         "int32",
            "col_idx":         "int16",
        })
    out_path = OUTPUT_DIR / f"{Path(release_file).stem}.parquet"
    df.to_parquet(out_path, index=False, compression="zstd")
    return out_path


def main() -> None:
    if not BRH_SOURCE.exists():
        sys.exit(f"BRH source folder not found: {BRH_SOURCE}")

    releases = discover_releases()
    if not releases:
        sys.exit(f"No matching .xlsx releases in {BRH_SOURCE} for "
                 f"{YEAR_MIN}-{YEAR_MAX}.")

    print(f"Audit formula extractor — Day 3 (xlsx only)", flush=True)
    print(f"Source: {BRH_SOURCE}", flush=True)
    print(f"Output: {OUTPUT_DIR}", flush=True)
    print(f"Scope: {len(releases)} xlsx releases\n", flush=True)

    total_formulas = 0
    t_start = time.monotonic()

    for i, (path, year, quarter) in enumerate(releases, 1):
        t0 = time.monotonic()
        try:
            df = extract_release_formulas(path, year, quarter)
        except Exception as e:
            print(f"  [{i:2d}/{len(releases)}] {path.name}: "
                  f"FAILED ({type(e).__name__}: {e})", flush=True)
            continue
        n = len(df)
        out_path = write_parquet(df, path.name)
        size_kb = out_path.stat().st_size / 1024
        elapsed = time.monotonic() - t0
        print(f"  [{i:2d}/{len(releases)}] {path.name:25s} "
              f"{n:>7,} formulas   {size_kb:6.1f} KB  ({elapsed:5.1f}s)",
              flush=True)
        total_formulas += n

    t_total = time.monotonic() - t_start
    print(f"\nDone. {total_formulas:,} formulas extracted across "
          f"{len(releases)} releases in {t_total:.1f}s.", flush=True)


if __name__ == "__main__":
    main()
