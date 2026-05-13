"""Build haiti-data.json for the Banking Sector Monitor dashboard.

Reads brh-dashboard/output/Financial_Monitoring_2025-12.xlsx and emits the JSON
shape consumed by data.js / monitor-utils.js.

FX trajectory (HTG per USD, end-of-quarter) is fetched live from IMF SDMX
(ER dataset, key HTI.XDC_USD.EOP_RT.Q) and cached on disk. If the fetch
fails, falls back to FX_SERIES_FALLBACK below.

Run from this folder:  python build_haiti_data.py
"""
from __future__ import annotations

import datetime as _dt
import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
XLSX = ROOT.parent / "FM Test" / "brh-dashboard" / "output" / "Financial_Monitoring_2025-12.xlsx"
OUT = ROOT / "haiti-data.json"

ACTIVE_BANKS = [
    {"ticker": "UNIBK",     "name": "Unibank",                       "color": "#1c3a5e"},
    {"ticker": "SOGEBK",    "name": "Sogebank",                      "color": "#7a4a1f"},
    {"ticker": "BNC",       "name": "Banque Nationale de Crédit",    "color": "#c19b5a"},
    {"ticker": "CAPITALBK", "name": "Capital Bank",                  "color": "#4d6b48"},
    {"ticker": "BUH",       "name": "Banque de l'Union Haïtienne",   "color": "#7d3c4a"},
    {"ticker": "SOGEBL",    "name": "Sogebel",                       "color": "#3a5e6b"},
    {"ticker": "BPH",       "name": "Banque Populaire Haïtienne",    "color": "#8a7a5e"},
]
LEGACY_BANKS = [
    {"ticker": "CBNA", "name": "Citibank N.A. (exited 2024)"},
]

INDICATOR_SHEETS = [
    "capital_to_assets", "car",
    "liquidity_to_assets", "liquidity_to_deposits",
    "total_assets", "net_loans", "gross_loans", "total_deposits",
    "shareholder_equity", "regulatory_capital", "rwa",
    "share_of_system_assets", "share_of_loans_plus_deposits", "share_of_deposits",
    "npl_ratio", "provision_coverage", "net_npl_to_equity",
    "roa", "roe",
]

# FX cache: refresh from IMF if older than this.
_FX_CACHE = Path(__file__).resolve().parent / ".fx_cache.json"
_FX_CACHE_MAX_AGE_DAYS = 7


def fetch_fx_from_imf() -> dict[str, float] | None:
    """Quarterly end-of-period HTG/USD from IMF SDMX. Returns {YYYY-MM-DD: rate} on
    quarter-end dates, or None if the fetch fails. Cached for a week.

    Uses the same SDMX client pattern as fetch_imf_panel.fetch_imf_data.
    """
    if _FX_CACHE.exists():
        age_days = (_dt.datetime.now().timestamp() - _FX_CACHE.stat().st_mtime) / 86400
        if age_days < _FX_CACHE_MAX_AGE_DAYS:
            print(f"  FX cache hit ({age_days:.1f}d old)")
            return json.loads(_FX_CACHE.read_text())

    try:
        import sdmx
    except ImportError:
        print("  ! sdmx not installed — pip install sdmx1 to enable IMF fetch")
        return None

    try:
        client = sdmx.Client("IMF_DATA")
        msg = client.data(
            "ER",
            key="HTI.XDC_USD.EOP_RT.Q",
            params={"startPeriod": "2014-Q4", "endPeriod": "2026-Q1"},
        )
        df = sdmx.to_pandas(msg).reset_index()[["TIME_PERIOD", "value"]]
    except Exception as e:
        print(f"  ! IMF fetch failed: {type(e).__name__}: {e}")
        return None

    # IMF reports e.g. "2025-Q4" → quarter-end date "2025-12-31"
    out: dict[str, float] = {}
    for _, row in df.iterrows():
        period = str(row["TIME_PERIOD"])  # e.g. "2025-Q4"
        try:
            year_s, q_s = period.split("-Q")
            year, q = int(year_s), int(q_s)
        except ValueError:
            continue
        last_month = q * 3
        # last day of month
        if last_month == 12:
            day = 31
        else:
            first_next = _dt.date(year, last_month + 1, 1)
            day = (first_next - _dt.timedelta(days=1)).day
        out[f"{year:04d}-{last_month:02d}-{day:02d}"] = float(row["value"])

    _FX_CACHE.write_text(json.dumps(out, indent=2))
    print(f"  fetched {len(out)} quarterly FX obs from IMF, cached")
    return out


# Fallback HTG/USD trajectory baked into the original handoff (used only if IMF fetch fails).
FX_SERIES_FALLBACK = {
    "2015-03-31": 49.17, "2015-06-30": 52.42, "2015-09-30": 55.67, "2015-12-31": 58.92,
    "2016-03-31": 61.40, "2016-06-30": 62.20, "2016-09-30": 63.50, "2016-12-31": 66.07,
    "2017-03-31": 67.50, "2017-06-30": 65.80, "2017-09-30": 64.20, "2017-12-31": 63.45,
    "2018-03-31": 64.30, "2018-06-30": 65.90, "2018-09-30": 66.50, "2018-12-31": 76.45,
    "2019-03-31": 81.20, "2019-06-30": 87.50, "2019-09-30": 92.20, "2019-12-31": 94.65,
    "2020-03-31": 95.40, "2020-06-30": 110.50, "2020-09-30": 116.10, "2020-12-31": 65.70,
    "2021-03-31": 75.10, "2021-06-30": 92.40, "2021-09-30": 96.50, "2021-12-31": 98.97,
    "2022-03-31": 105.80, "2022-06-30": 119.40, "2022-09-30": 122.20, "2022-12-31": 145.30,
    "2023-03-31": 152.10, "2023-06-30": 144.60, "2023-09-30": 137.20, "2023-12-31": 132.85,
    "2024-03-31": 132.40, "2024-06-30": 132.20, "2024-09-30": 131.50, "2024-12-31": 131.10,
    "2025-03-31": 130.80, "2025-06-30": 130.50, "2025-09-30": 130.30, "2025-12-31": 130.08,
}

# Excel uses a non-ASCII variant of "SYSTÈME" as the column header — match by prefix.
SYSTEM_PREFIX = "SYST"

# The Excel carries history back to ~2001; the dashboard's 10y window starts 2015Q1.
MIN_DATE = "2015-01-01"


def _date_str(value) -> str | None:
    if value is None:
        return None
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value)


def _read_indicator(ws) -> tuple[list[str], dict[str, dict[str, float]]]:
    """Return (column tickers, {date: {ticker: value}}) for an indicator sheet."""
    rows = list(ws.iter_rows(values_only=True))
    header_idx = next(i for i, r in enumerate(rows) if r and r[0] == "Date")
    headers = rows[header_idx]
    tickers = [str(h) if h else None for h in headers[1:]]

    out: dict[str, dict[str, float]] = {}
    for r in rows[header_idx + 1 :]:
        if not r or r[0] is None:
            continue
        date = _date_str(r[0])
        if date is None or date < MIN_DATE:
            continue
        bucket: dict[str, float] = {}
        for col, val in zip(tickers, r[1:]):
            if col is None or val is None:
                continue
            key = "SYSTEM" if col.startswith(SYSTEM_PREFIX) else col
            if key in {b["ticker"] for b in ACTIVE_BANKS} or key == "SYSTEM":
                bucket[key] = float(val)
        if bucket:
            out[date] = bucket
    return tickers, out


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    values: dict[str, dict[str, dict[str, float]]] = {}
    all_dates: set[str] = set()
    for sheet in INDICATOR_SHEETS:
        if sheet not in wb.sheetnames:
            print(f"  ! missing sheet: {sheet}")
            continue
        _, by_date = _read_indicator(wb[sheet])
        values[sheet] = by_date
        all_dates.update(by_date.keys())

    # System-weighted overrides
    system_weighted: dict[str, dict[str, float]] = {}
    for sheet, key in (("system_npl_weighted", "npl"), ("system_car_weighted", "car")):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        header_idx = next(i for i, r in enumerate(rows) if r and r[0] == "Date")
        for r in rows[header_idx + 1 :]:
            if not r or r[0] is None or r[1] is None:
                continue
            date = _date_str(r[0])
            if date is None or date < MIN_DATE:
                continue
            system_weighted.setdefault(date, {})[key] = float(r[1])

    dates = sorted(all_dates)

    # FX: try IMF, fall back to baked-in series.
    print("Fetching FX from IMF SDMX (HTI.XDC_USD.EOP_RT.Q)...")
    fx_source = fetch_fx_from_imf()
    if fx_source is None:
        print("  using fallback FX series")
        fx_source = FX_SERIES_FALLBACK

    fx: dict[str, float] = {}
    for d in dates:
        if d in fx_source:
            fx[d] = fx_source[d]
            continue
        prior = [k for k in fx_source if k <= d]
        if not prior:
            raise RuntimeError(f"no FX rate available at or before {d}")
        ref = max(prior)
        print(f"  ! no FX for {d} — carrying forward from {ref} ({fx_source[ref]:.2f})")
        fx[d] = fx_source[ref]

    payload = {
        "banks": ACTIVE_BANKS,
        "legacy": LEGACY_BANKS,
        "dates": dates,
        "values": values,
        "systemWeighted": system_weighted,
        "fx": fx,
    }

    OUT.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    size_kb = OUT.stat().st_size / 1024
    print(f"wrote {OUT.name}: {len(dates)} quarters, {len(values)} indicators, {size_kb:.1f} KB")


if __name__ == "__main__":
    main()
